import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dataclasses import dataclass
import csv
from datetime import date

OFFICE_HOURS_STR = "Office Hours"

@dataclass
class InstructorInfo:
    name: str
    ext: str
    email: str
    office_location: str

@dataclass
class CourseInfo:
    name_and_section: str 
    ty: str
    days: list[str]
    start_time: tuple[int, int]
    end_time: tuple[int, int]
    instructor: InstructorInfo
    location: str

# function splits time apart 
def parse_time(time: str) -> tuple[int, int]:
    parts_of_time = time.replace(":", " ").split(" ")

    hour = int(parts_of_time[0])

    # convert to 24 hour time for the excel file AM/PM
    if parts_of_time[2] == "PM" and hour != 12:
        hour += 12
    
    min = int(parts_of_time[1])
    return (hour, min)

# reuses data from previous calls to the get_instructor_info function
instructor_info_cache = {}
def get_instructor_info(name: str, info_url: str) -> InstructorInfo:

    if instructor_info_cache.get(info_url):
        return instructor_info_cache[info_url]
    
    raw_html = requests.get(info_url).text
    soup = BeautifulSoup(raw_html, "html.parser")

    alias = soup.find("span", {"class": "alias"})
    phone_number_maybe = alias.find("a")
    ext = ""

    if phone_number_maybe is not None: # if it exists
        ext = f"x6-{phone_number_maybe.text.split('.')[-1]}"

    alias_parts = alias.text.split("*")
    office_location = alias_parts[0].strip()

    username = alias_parts[1].strip()
    email = ""

    if username != "": # if it exists
        email = f"{username}@calpoly.edu"


    instructor_info = InstructorInfo(
        office_location = office_location,
        email = email,
        ext = ext,
        name = name
    )

    instructor_info_cache[info_url] = instructor_info

    return instructor_info
    

def get_course_list() -> list[CourseInfo]:
    # this request url comes from schedules.calpoly.edu
    # it is a get request
    raw_html = requests.get("https://schedules.calpoly.edu/subject_IME_next.htm").text
    # uses the library, takes the text and converts to a structure that can be searched 
    soup = BeautifulSoup(raw_html, "html.parser")
    # tr refers to the rows of data on the website
    # skip the first row because that is the row with the headers and not the data that we want
    # course_rows is all the rows
    # course_row is the individual row
    course_rows = soup.find_all("tr")[1::]

    courses: list[CourseInfo] = []

    # for each row in the total rows
    for course_row in course_rows:
        # in each row in all the rows, find td > class > courseName, because there is an a for the course name, add that too, then get the course name text
        course_start_time = course_row.find("td", {"class": "startTime"}).text
        course_instructor = course_row.find("td", {"class": "personName"}).text

        if course_start_time == "\xa0" or course_instructor == "\xa0":
            continue # only get things with start times (dont want classes that don't go on schedule)

        course_name = course_row.find("td", {"class": "courseName"}).find("a").text
        course_section = course_row.find("td", {"class": "courseSection"}).text
        course_name_and_section = f"{course_name}-{course_section}"
        course_type = course_row.find("td", {"class": "courseType"}).text
        course_days = [*course_row.find("td", {"class": "courseDays"}).text] # [*~~~] splits up all thing in string as separate item in list
        course_end_time = course_row.find("td", {"class": "endTime"}).text
        instructor_name = course_row.find("td", {"class": "personName"}).find("a").get("title")
        instructor_url = course_row.find("td", {"class": "personName"}).find("a").get("href")
        course_location = course_row.find("td", {"class": "location"}).find("a").text

        course = CourseInfo(
            name_and_section = course_name_and_section,
            ty = course_type,
            days = course_days,
            start_time = parse_time(course_start_time),
            end_time = parse_time(course_end_time),
            instructor = get_instructor_info(instructor_name, f"https://schedules.calpoly.edu/{instructor_url}"),
            location = course_location
        )
        courses.append(course)

    return courses

def group_courses_by(courses: list[CourseInfo], key_path: str) -> dict[str, list[CourseInfo]]:
    output = {}
    keys = key_path.split(".")

    for course in courses:
        course_dict = course.__dict__

        grouping_value = course_dict
        for key in keys:
            if hasattr(grouping_value[key], "__dict__"):
                grouping_value = grouping_value[key].__dict__
            else:
                grouping_value = grouping_value[key]

        if output.get(grouping_value) is not None:
            output[grouping_value].append(course)
        else:
            output[grouping_value] = [course]
    return output


def time_to_row(time: tuple[int, int]) -> int:
    hour = time[0]
    row =  (hour - 7) * 2 + 3

    if time[1] == 40:
        row += 1

    return row


# map of days to columns
day_to_column = {
    "M" : "B", # when day is M, the column is B 
    "T" : "C",
    "W" : "D",
    "R" : "E",
    "F" : "F",
}


def add_instructor_page(workbook: Workbook, courses: list[CourseInfo]):
    template = workbook["Template"] # this refers to the worksheet named template
    # create new worksheet 
    instructor = courses[0].instructor
    new_sheet = workbook.copy_worksheet(template) # create new sheet 
    new_sheet.title = instructor.name

    # add the instructor name 
    new_sheet['B1'] = instructor.name
    new_sheet['D1'] = instructor.ext
    new_sheet['E1'] = f"Office: {instructor.office_location}"
    new_sheet['G1'] = instructor.email

    for course in courses:

        # find start and end row
        start_row = time_to_row(course.start_time)
        end_row = time_to_row(course.end_time) - 1
      
        # merge appropriate cells
        for day in course.days:
            column = day_to_column[day]
            new_sheet.merge_cells(f"{column}{start_row}:{column}{end_row}")
            target_cell = new_sheet[f"{column}{start_row}"] 
            if end_row - start_row < 3 or course.name_and_section == OFFICE_HOURS_STR:
                target_cell.value = f"{course.name_and_section} {course.ty}\n{course.location}"
            else: 
                target_cell.value = f"{course.name_and_section}\n{course.ty}\n{course.location}"
            target_cell.font = Font(bold=True)
            target_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if course.name_and_section == OFFICE_HOURS_STR:
                target_cell.fill = PatternFill("solid", fgColor="ffff00")
            else:
                target_cell.fill = PatternFill("solid", fgColor="c6e0b4")
                


    # use first digits of time for index * 2 + offset
    # if minutes = 10 add 0 to index. If minutes is 40 add 1 to index
    # merge f"{day_map[day]{start_index}:{day_map[day]}{end_index}"
    # new_sheet[f"{day_map[day]{start_index}"] = course.name_and_section


def main():
    courses = get_course_list()
    courses_by_instructor = group_courses_by(courses, "instructor.name")

    # import instructor schedules template
    excel_doc = openpyxl.load_workbook("template.xlsx")

    with open("office_hours.csv") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        
        # skip header of file
        next(csv_reader)
        for row in csv_reader:
            name, days_str, start_time, end_time, location = row
            course_info = CourseInfo(
                name_and_section = OFFICE_HOURS_STR,
                start_time = parse_time(start_time),
                end_time = parse_time(end_time),
                days = [*days_str],
                location = location,
                ty = "",
                instructor = None
            )
            courses_by_instructor[name].append(course_info)

    for taught_courses in courses_by_instructor.values():
        add_instructor_page(excel_doc, taught_courses)
    
    today = date.today()
    excel_doc.save(f"Instructor_Schedules_{today.strftime('%m-%d-%y')}.xlsx")


# name is when file is imported 
# if you are running this file, call this function 
if __name__ == "__main__":
    main()
